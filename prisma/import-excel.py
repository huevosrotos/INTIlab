#!/usr/bin/env python3
"""
Lee el Excel del droguero y genera un JSON con las sustancias categorizadas.
Salida: prisma/seed-data.json
"""

import openpyxl
import json
import re
import os

EXCEL_PATH = "/home/z/my-project/upload/Stock drogas-CONTROL II.xlsx"
OUTPUT_PATH = "/home/z/my-project/prisma/seed-data.json"

# ============================================================
# Función de categorización automática
# ============================================================
def categorize(name: str, tipo: str) -> list:
    """Devuelve lista de clases químicas basada en el nombre."""
    text = f"{name} {tipo}".lower()
    classes = []

    # Indicadores de pH
    ph_indicators = ["fenolftaleína", "fenolftaleina", "heliantina", "naranja de metilo",
                     "azul de bromofenol", "rojo de metilo", "púrpura de bromocresol",
                     "verde de bromocresol", "timolftaleína", "tropeolina", "alizarina",
                     "curcumina", "rojo fenol", "rojo neutral", "rojo de cresol",
                     "amarillo de metilo", "verde de malaquita", "naranja de xilenol"]
    for ind in ph_indicators:
        if ind in text:
            classes.append("INDICADOR_PH")
            break

    # Indicadores redox
    redox_indicators = ["difenilamina", "ferroína", "ferroina", "azul de metileno",
                        "dicromato", "yoduro de almidón", "almidón", "ortofenantrolina",
                        "fenantrolina", "difeni", "carbazón", "ditiocarbamato"]
    for ind in redox_indicators:
        if ind in text:
            classes.append("INDICADOR_REDOX")
            break

    # Colorantes
    dyes = ["fucsina", "violeta", "cristal", "azul de", "rojo de", "verde de",
            "amaranto", "eosina", "eritrosina", "nigrosina", "safranina",
            "fluoresceína", "fluoresceina", "rojo neutro", "azul de toluidina",
            "verde de janus", "ponceau", "coomassie", "orange g", "orceína",
            "hematoxilina", "carbón", "carmin", "iodine green", "pontamine",
            "solochrome", "murexida", "hofman"]
    for dye in dyes:
        if dye in text:
            classes.append("COLORANTE")
            break

    # Solventes
    solvents = ["acetona", "etanol", "metanol", "isopropanol", "tolueno", "xileno",
                 "éter", "eter", "cloroformo", "dicloro", "acetonitrilo", "hexano",
                 "heptano", "petróleo", "petroleo", "benceno", "acetato de etilo",
                 "piridina", "tetrahidrofurano", "dimetil", "dmso",
                 "alcohol"]
    for sol in solvents:
        if sol in text:
            classes.append("SOLVENTE")
            break

    # Ácidos
    if "ácido" in text or "acido" in text or text.strip().startswith("acido") or " buffer ph" in text:
        if "buffer" not in text:
            classes.append("ACIDO")
    # Bases
    bases = ["hidróxido", "hidroxido", "óxido de", "oxido de", "amoníaco", "amoniaco",
             "carbonato de", "bicarbonato", "amina", "amina,", "etanolamina",
             "tri", "piridina", "cal hidratada", "cal viva"]
    for b in bases:
        if b in text and "ácido" not in text and "acido" not in text:
            classes.append("BASE")
            break

    # Sales (si no es ácido ni base pero tiene "de" o "cloruro" o "sulfato")
    salts = ["cloruro de", "sulfato de", "nitrato de", "carbonato de", "fosfato de",
             "acetato de", "oxalato de", "citrato de", "tartrato de", "cianuro de",
             "bromuro de", "yoduro de", "fluoruro de", "permanganato de", "dicromato de",
             "tiocianato de", "nitrato", "cloruro", "sulfato", "carbonato", "fosfato"]
    is_acid = "ACIDO" in classes
    is_base = "BASE" in classes
    if not is_acid and not is_base:
        for s in salts:
            if s in text:
                classes.append("SAL")
                break

    # Peligrosidad
    dangerous = ["mercurio", "arsénico", "arsenico", "cianuro", "plomo", "cadmio",
                 "cromo", "níquel", "niquel", "berilio", "talio", "bencidina"]
    for d in dangerous:
        if d in text:
            classes.append("TOXICO")
            break

    if "flamable" in text or "inflamable" in text or "éter" in text or "eter diet" in text:
        classes.append("INFLAMABLE")
    if "explos" in text or "peróxido" in text or "peroxido" in text:
        classes.append("EXPLOSIVO")
    if "corros" in text or "ácido sulfúrico" in text or "ácido clorhídrico" in text or "ácido nítrico" in text:
        classes.append("CORROSIVO")
    if "oxidante" in text or "comburente" in text or "permanganato" in text or "dicromato" in text:
        classes.append("COMBURENTE")

    # Naturaleza (orgánico/inorgánico)
    # Metales → inorgánico
    metals = ["cloruro de", "sulfato de", "nitrato de", "óxido de", "oxido de",
              "carbonato de", "fosfato de", "mercurio", "plomo", "cobre", "hierro",
              "aluminio", "calcio", "sodio", "potasio", "magnesio", "zinc", "níquel",
              "niquel", "cromo", "cobalto", "manganeso", "bario", "estroncio",
              "cesio", "litio", "plata", "oro", "platino", "paladio", "rodio",
              "estaño", "antimonio", "bismuto", "cadmio", "alumin", "metalico", "metálico"]
    is_metal = False
    for m in metals:
        if m in text:
            classes.append("COMPUESTO_METALICO")
            is_metal = True
            break

    # Metales puros
    pure_metals = ["cobre metálico", "aluminio metal", "calcio metál", "sodio metál",
                   "estaño metál", "zinc metál", "magnesio metál"]
    for m in pure_metals:
        if m in text:
            classes.append("METAL")
            is_metal = True
            break

    # Determinar orgánico/inorgánico
    organic_markers = ["benz", "fenol", "anilina", "tolueno", "xileno", "benceno",
                       "naftalen", "piridina", "antraceno", "fenantrolina",
                       "ácido acético", "acido acetico", "acetato", "acetona",
                       "etanol", "metanol", "éter", "formaldehído", "formol",
                       "glucosa", "fructosa", "sacarosa", "almidón", "celulosa",
                       "proteína", "albumina", "gelatina", "aceite", "cera",
                       "cloroformo", "tetracloruro", "carbón activo", "carmin",
                       "fucsina", "violeta", "safranina", "hematoxilina"]
    is_organic = False
    for org in organic_markers:
        if org in text:
            is_organic = True
            break

    if is_organic:
        classes.append("ORGANICO")
    else:
        classes.append("INORGANICO")

    # Biológicos
    if "glucosa" in text or "fructosa" in text or "galactosa" in text or "sacarosa" in text:
        classes.append("CARBOHIDRATO")
    if "almidón" in text or "almidon" in text or "celulosa" in text or "carboximetil" in text:
        classes.append("CARBOHIDRATO")
    if "albumina" in text or "proteína" in text or "proteina" in text or "gelatina" in text:
        classes.append("PROTEINA")
    if "aceite" in text or "lipido" in text or "lípido" in text or "colesterol" in text:
        classes.append("LIPIDO")
    if "polímero" in text or "polimero" in text or "carboximetilcelulosa" in text:
        classes.append("POLIMERO")
    if "surfactante" in text or "detergente" in text or "tween" in text or "tritón" in text:
        classes.append("SURFACTANTE")

    # Buffer
    if "buffer" in text or "buffer ph" in text:
        classes.append("BUFFER")

    # Si no tiene nada, marcar como OTRO
    if not classes:
        classes.append("OTRO")

    # Deduplicar manteniendo orden
    seen = set()
    unique = []
    for c in classes:
        if c not in seen:
            seen.add(c)
            unique.append(c)
    return unique


# ============================================================
# Generar código nuevo: TIPO-ARMARIO-ESTANTE-CORR
# ============================================================
def generate_code(clases: list, armario: int, estante: str, corr: int) -> str:
    """Genera el código nuevo. TIPO = primeras letras de la clase principal."""
    # Mapeo de clase → prefijo
    prefixes = {
        "ACIDO": "AC",
        "BASE": "BA",
        "SAL": "SA",
        "SOLVENTE": "SO",
        "INDICADOR_PH": "IP",
        "INDICADOR_REDOX": "IR",
        "COLORANTE": "CO",
        "REACTIVO_ANALITICO": "RA",
        "BUFFER": "BU",
        "INFLAMABLE": "IN",
        "TOXICO": "TO",
        "CORROSIVO": "CO",
        "COMBURENTE": "CB",
        "ORGANICO": "OR",
        "INORGANICO": "IO",
        "METAL": "ME",
        "COMPUESTO_METALICO": "CM",
        "CARBOHIDRATO": "CA",
        "PROTEINA": "PR",
        "LIPIDO": "LI",
        "POLIMERO": "PO",
        "SURFACTANTE": "SU",
        "OTRO": "XX",
    }
    # Tomar la primera clase que no sea naturaleza
    primary = "OTRO"
    priority = ["ACIDO", "BASE", "SOLVENTE", "INDICADOR_PH", "INDICADOR_REDOX",
                "COLORANTE", "SAL", "BUFFER", "METAL", "COMPUESTO_METALICO",
                "CARBOHIDRATO", "PROTEINA", "LIPIDO", "POLIMERO", "SURFACTANTE",
                "INFLAMABLE", "TOXICO", "CORROSIVO", "ORGANICO", "INORGANICO"]
    for p in priority:
        if p in clases:
            primary = p
            break
    prefix = prefixes.get(primary, "XX")
    return f"{prefix}-{armario}{estante}-{corr:03d}"


# ============================================================
# Determinar estado físico y unidad
# ============================================================
def detect_state_and_unit(name: str, stock: str) -> tuple:
    text = f"{name} {stock}".lower()
    # Estado
    if any(x in text for x in ["litro", " ml", "l ", "líquido", "liquido", "solvente"]):
        state = "LIQUIDO"
    elif any(x in text for x in [" g", "g ", "gramo", "polvo", "sólido", "solido", "cristal"]):
        state = "SOLIDO"
    elif any(x in text for x in ["gas", "gaseoso"]):
        state = "GAS"
    else:
        state = "SOLIDO"  # default

    # Unidad
    if "litro" in text or "l " in text or text.rstrip().endswith("l"):
        unit = "L"
    elif "ml" in text:
        unit = "mL"
    elif "g " in text or "gramo" in text or text.rstrip().endswith("g"):
        unit = "g"
    elif "kg" in text:
        unit = "kg"
    else:
        unit = "u"  # unidades (sobres, etc.)

    return state, unit


# ============================================================
# Determinar depósito según la hoja
# ============================================================
def detect_warehouse(sheet_name: str) -> str:
    s = sheet_name.lower().strip()
    if "hplc" in s:
        return "Solventes HPLC"
    if "solvent" in s:
        return "Solventes"
    if "acid" in s:
        return "Ácidos"
    if "consum" in s:
        return "Consumibles"
    return "Depósito Central / Droguero"


# ============================================================
# Mapear nombre (B) + tipo (C) a nombre químico limpio
# ============================================================
# El Excel tiene las columnas B y C que a veces están invertidas:
# - Caso normal: B="Ácido sulfúrico", C=None → "Ácido sulfúrico"
# - Sal con anión en C: B="zinc", C="Sulfato de" → "Sulfato de zinc"
# - Sal con catión en C: B="Cloruro de", C="mercurio" → "Cloruro de mercurio"
# - Ácido: B="Ácido", C="perídico" → "Ácido periódico"

# Aniones que pueden ir en C (como "Sulfato de zinc")
ANION_PREFIXES = [
    "cloruro", "sulfato", "nitrato", "carbonato", "fosfato", "acetato",
    "oxalato", "citrato", "tartrato", "cianuro", "bromuro", "yoduro",
    "fluoruro", "permanganato", "dicromato", "tiocianato", "oxido", "óxido",
    "hidróxido", "hidroxido", "nitrato", "clorato", "perclorato", "sulfuro",
    "nitrito", "borato", "silicato", "tiosulfato", "bicarbonato", "bisulfato",
    "cromato", "molibdato", "tungstato", "vanadato", "arseniato", "arsenito",
    "selenito", "telurito", "permanganato", "dihidrato", "hidrato",
]

def normalize_hydrates(name: str) -> str:
    """Normaliza los nombres de compuestos hidratados.
    Convierte 'Acetato dihidrato de Zinc' → 'Acetato de Zinc (dihidrato)'
    Convierte 'Cloruro de (anhidro) de Calcio' → 'Cloruro de Calcio (anhidro)'
    """
    import re

    # Patrones de hidratación que pueden aparecer
    hydrate_patterns = [
        # "dihidrato de Zinc" → mover "dihidrato" al final entre paréntesis
        (r'\s+(monohidrato|dihidrato|trihidrato|tetrahidrato|pentahidrato|hexahidrato|heptahidrato|octahidrato|decahidrato|hidrato)\s+de\s+', ' de '),
        # "(anhidro) de Sodio" → "de Sodio (anhidro)"
        (r'\s*\((anhidro|anhidro cristal|anhidra)\)\s*de\s*', ' de '),
        # "de(anhidro) de" → "de ... (anhidro)"
        (r'\s*de\s*\((anhidro|monohidrato|dihidrato|trihidrato|tetrahidrato|pentahidrato|hexahidrato|heptahidrato|decahidrato|6-Hidrato|hidrato)\)\s*de\s*', ' de '),
        # "6-Hidrato" → "(hexahidrato)"
        (r'\s*\(?\s*(\d+-?[Hh]idrato)\s*\)?\s*', ''),
        # "(heptahidrato)" ya en paréntesis - dejarlo
        # "anhidro" suelto al final → "(anhidro)"
        (r'\s+anhidro$', ''),
    ]

    # Extraer el tipo de hidratación si está presente
    hydrate_type = None
    remaining = name

    # Buscar patrones de hidratación
    hydrate_keywords = [
        'monohidrato', 'dihidrato', 'trihidrato', 'tetrahidrato',
        'pentahidrato', 'hexahidrato', 'heptahidrato', 'octahidrato',
        'decahidrato', 'anhidro', 'anhidra', 'hidrato',
        '6-hidrato', '6-Hidrato', 'clorhidrato', 'diclorhidrato',
    ]

    # Mapeo de números a prefijos griegos
    num_to_prefix = {
        '1': 'monohidrato', '2': 'dihidrato', '3': 'trihidrato',
        '4': 'tetrahidrato', '5': 'pentahidrato', '6': 'hexahidrato',
        '7': 'heptahidrato', '8': 'octahidrato', '9': 'nonahidrato',
        '10': 'decahidrato',
    }

    # Caso: "X (dihidrato)" → ya está bien, solo limpiar
    m = re.search(r'\((monohidrato|dihidrato|trihidrato|tetrahidrato|pentahidrato|hexahidrato|heptahidrato|octahidrato|decahidrato|anhidro|anhidro cristal)\)', remaining, re.IGNORECASE)
    if m:
        hydrate_type = m.group(1).lower()
        # Quitar el paréntesis del nombre, lo volveremos a agregar al final
        remaining = remaining[:m.start()] + remaining[m.end():]
        remaining = re.sub(r'\s+', ' ', remaining).strip().rstrip('de').strip()
    else:
        # Caso: "dihidrato de Zinc" → extraer "dihidrato"
        for kw in hydrate_keywords:
            if kw.lower() in remaining.lower() and kw.lower() not in ['clorhidrato', 'diclorhidrato']:
                # No confundir clorhidrato con hidrato
                pattern = re.compile(re.escape(kw), re.IGNORECASE)
                # Verificar que no sea parte de "clorhidrato"
                idx = remaining.lower().find(kw.lower())
                if idx > 0 and remaining[idx-1:idx].lower() == 'o' and 'clor' in remaining[max(0,idx-5):idx].lower():
                    continue  # Es clorhidrato, no hidrato
                hydrate_type = kw.lower()
                # Remover del nombre
                remaining = pattern.sub('', remaining)
                break

        # Caso: "6-Hidrato" o "6-Hidrato)"
        m2 = re.search(r'(\d+-?[Hh]idrato)', remaining)
        if m2 and not hydrate_type:
            num = m2.group(1).split('-')[0].lower()
            hydrate_type = num_to_prefix.get(num, f'{num}-hidrato')
            remaining = remaining[:m2.start()] + remaining[m2.end():]

    # Limpiar el nombre restante
    remaining = re.sub(r'\s+', ' ', remaining).strip()
    # Quitar "de " sobrante al final
    remaining = re.sub(r'\s+de$', '', remaining).strip()
    # Quitar "de de" → "de" (se repite por el split)
    remaining = re.sub(r'\bde\s+de\b', 'de', remaining).strip()
    remaining = re.sub(r'\bde\s+de\b', 'de', remaining).strip()
    # Quitar paréntesis vacíos
    remaining = re.sub(r'\(\s*\)', '', remaining).strip()
    # Quitar paréntesis con números sueltos como "(6-)"
    remaining = re.sub(r'\(\d+-?\)', '', remaining).strip()
    # Quitar comas sueltas
    remaining = re.sub(r',\s*$', '', remaining).strip()
    # Quitar "Clor" suelto (que viene de clorhidrato mal cortado)
    remaining = re.sub(r'\s+Clor$', '', remaining).strip()
    remaining = re.sub(r'\s+Clor\s+', ' ', remaining).strip()
    # Quitar "Dichlor" suelto
    remaining = re.sub(r'\s+Dichlor\s+', ' ', remaining).strip()
    remaining = re.sub(r'\s+Dichlor$', '', remaining).strip()

    # Si encontramos tipo de hidratación, agregarlo al final entre paréntesis
    if hydrate_type:
        # Normalizar "anhidro cristal" → "anhidro"
        if 'anhidro' in hydrate_type:
            hydrate_type = 'anhidro'
        return f"{remaining} ({hydrate_type})"

    return remaining


def clean_name(name_b: str, tipo_c: str) -> str:
    """Combina B y C para obtener el nombre químico correcto."""
    b = (name_b or "").strip()
    c = (tipo_c or "").strip()

    # Si no hay C, devolver B con capitalización correcta
    if not c:
        if not b:
            return ""
        return b[0].upper() + b[1:] if b else b

    # Si no hay B, devolver C
    if not b:
        return c

    b_lower = b.lower()
    c_lower = c.lower()

    # Caso 1: C es un anión/prefijo que va antes del catión
    # ej: B="zinc", C="Sulfato de" → "Sulfato de zinc"
    # ej: B="plata", C="Nitrato de" → "Nitrato de plata"
    for anion in ANION_PREFIXES:
        if c_lower.startswith(anion) or c_lower == anion:
            # Capitalizar el catión
            cat = b[0].upper() + b[1:] if b else b
            # Si C termina con "de", armar "C catión"
            if c_lower.endswith("de") or c_lower.endswith("de "):
                return f"{c} {cat}".strip()
            # Si C no tiene "de", agregarlo
            return f"{c} de {cat}".strip()

    # Caso 2: B es un anión/prefijo que va antes del catión
    # ej: B="Cloruro de", C="mercurio" → "Cloruro de mercurio"
    # ej: B="Nitrato de", C="plata" → "Nitrato de plata"
    for anion in ANION_PREFIXES:
        if b_lower.startswith(anion):
            cat = c[0].upper() + c[1:] if c else c
            if b_lower.endswith("de") or b_lower.endswith("de "):
                return f"{b} {cat}".strip()
            return f"{b} de {cat}".strip()

    # Caso 3: B="Ácido", C="perídico" → "Ácido periódico"
    if b_lower in ["ácido", "acido"] and c:
        return f"Ácido {c}".strip()

    # Caso 4: C="Ácido", B="perídico" → "Ácido periódico"
    if c_lower in ["ácido", "acido"] and b:
        return f"Ácido {b}".strip()

    # Caso 5: Ambos son texto, concatenar "B C"
    # ej: B="Paladium", C="Nitrato Hidrato" → "Paladium Nitrato Hidrato"
    if b and c and not c_lower.startswith("de"):
        return f"{b} {c}".strip()

    # Default: solo B
    return b if b else c


# ============================================================
# Main
# ============================================================
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    drugs = {}  # codigo_viejo → {nombre, clases, lotes}
    lot_counter = {}  # codigo_viejo → contador de lotes

    for ws in wb.worksheets:
        sheet_name = ws.title
        warehouse = detect_warehouse(sheet_name)

        for i in range(2, ws.max_row + 1):
            row = [c.value for c in ws[i]]
            if not row or not row[0]:
                continue

            codigo = str(row[0]).strip()
            if not codigo or not codigo.isdigit():
                continue

            name_raw = str(row[1]).strip() if row[1] else ""
            tipo_raw = str(row[2]).strip() if row[2] else ""
            # Columna E = Cantidad/stock (índice 4)
            stock_raw = str(row[4]).strip() if row[4] else ""
            # Columna F = Calidad (índice 5)
            calidad = str(row[5]).strip() if row[5] else ""
            # Columna G = Marca (índice 6)
            marca = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            # Columna H = Ubicación/depósito (índice 7)
            ubic_raw = str(row[7]).strip() if len(row) > 7 and row[7] else ""

            # Determinar depósito: si la columna H tiene "DEPOSITO", usar el de la hoja
            # (la columna H a veces tiene info de ubicación física)
            lot_warehouse = warehouse
            if ubic_raw and ubic_raw.lower() not in ["deposito", ",", ".", ""]:
                lot_warehouse = ubic_raw  # usar la ubicación específica si existe

            # Limpiar nombre combinando B y C
            nombre = clean_name(name_raw, tipo_raw)
            # Normalizar: quitar espacios dobles, capitalizar primera letra
            nombre = re.sub(r"\s+", " ", nombre).strip()
            if nombre:
                nombre = nombre[0].upper() + nombre[1:]
            # Normalizar hidratos: "Acetato dihidrato de Zinc" → "Acetato de Zinc (dihidrato)"
            nombre = normalize_hydrates(nombre)
            if not nombre or len(nombre) < 2:
                continue

            # Categorizar usando el nombre completo
            clases = categorize(nombre, "")

            # Estado y unidad
            state, unit = detect_state_and_unit(nombre, stock_raw)

            # Extraer cantidad numérica del stock
            qty = 0
            if stock_raw:
                # Buscar número (puede tener coma decimal)
                m = re.search(r"(\d+(?:[.,]\d+)?)", stock_raw)
                if m:
                    try:
                        qty = float(m.group(1).replace(",", "."))
                    except:
                        qty = 0
            if qty == 0:
                qty = 1  # default

            # Limpiar marca (a veces tiene notas entre paréntesis)
            marca_clean = marca
            if marca:
                # Quitar paréntesis con notas (ej: "Biopack(baja 22/05/17)" → "Biopack")
                marca_clean = re.sub(r"\(.*?\)", "", marca).strip()
                if not marca_clean:
                    marca_clean = marca

            # Inicializar drug si no existe
            if codigo not in drugs:
                drugs[codigo] = {
                    "codigo_viejo": codigo,
                    "chemicalName": nombre,
                    "clases": clases,
                    "state": state,
                    "unit": unit,
                    "purity": calidad if calidad else None,
                    "warehouse": warehouse,
                    "lotes": []
                }
                lot_counter[codigo] = 0

            # Agregar lote (frasco)
            lot_counter[codigo] += 1
            drugs[codigo]["lotes"].append({
                "lotNumber": f"{codigo}-{lot_counter[codigo]}",
                "supplier": marca_clean if marca_clean else None,
                "initialQuantity": qty,
                "currentQuantity": qty,
                "purity": calidad if calidad else None,
            })

    # Generar códigos nuevos
    # Asignar armario y estante según orden de aparición
    armarios = [1, 2, 3, 4, 5, 6]
    estantes = ["A", "B", "C", "D"]
    armario_idx = 0
    estante_idx = 0
    corr_por_armario_estante = {}

    # Ordenar drogas por nombre para asignar códigos
    codigos_ordenados = sorted(drugs.keys(), key=lambda c: drugs[c]["chemicalName"].lower())

    output_drugs = []
    for codigo in codigos_ordenados:
        drug = drugs[codigo]
        # Asignar armario/estante (4 estantes por armario, ~50 drogas por estante)
        key = (armarios[armario_idx], estantes[estante_idx])
        if key not in corr_por_armario_estante:
            corr_por_armario_estante[key] = 0
        corr_por_armario_estante[key] += 1
        corr = corr_por_armario_estante[key]

        new_code = generate_code(drug["clases"], armarios[armario_idx], estantes[estante_idx], corr)

        drug["code"] = new_code
        drug["armario"] = armarios[armario_idx]
        drug["estante"] = estantes[estante_idx]
        drug["correlativo"] = corr

        output_drugs.append(drug)

        # Avanzar estante cada 50 drogas
        if corr >= 50:
            estante_idx += 1
            if estante_idx >= len(estantes):
                estante_idx = 0
                armario_idx += 1
                if armario_idx >= len(armarios):
                    armario_idx = 0  # wrap (no debería pasar con 297 drogas)

    # Guardar JSON
    output = {
        "warehouses": [
            {"name": "Depósito Central / Droguero", "code": "DEP-00", "type": "PRINCIPAL",
             "location": "Edificio A - Planta Baja"},
            {"name": "Solventes HPLC", "code": "DEP-01", "type": "SECUNDARIO",
             "location": "Droguero Central"},
            {"name": "Solventes", "code": "DEP-02", "type": "SECUNDARIO",
             "location": "Droguero Central"},
            {"name": "Ácidos", "code": "DEP-03", "type": "SECUNDARIO",
             "location": "Droguero Central"},
            {"name": "Consumibles", "code": "DEP-04", "type": "SECUNDARIO",
             "location": "Droguero Central"},
            {"name": "Lab-MEIPA", "code": "LAB-01", "type": "SECUNDARIO",
             "location": "Laboratorio MEIPA"},
            {"name": "Lab-SERVICIOS", "code": "LAB-02", "type": "SECUNDARIO",
             "location": "Laboratorio de Servicios"},
            {"name": "Lab-Microbiología", "code": "LAB-03", "type": "SECUNDARIO",
             "location": "Laboratorio de Microbiología"},
            {"name": "Lab-Cromatografía", "code": "LAB-04", "type": "SECUNDARIO",
             "location": "Laboratorio de Cromatografía"},
            {"name": "Lab-Bioprocesos", "code": "LAB-05", "type": "SECUNDARIO",
             "location": "Laboratorio de Bioprocesos"},
        ],
        "users": [
            {"email": "mdellavecchia@inti.gob.ar", "name": "MDV", "password": "3141",
             "role": "ADMIN"},
        ],
        "drugs": output_drugs,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"✓ Generado {OUTPUT_PATH}")
    print(f"  - {len(output_drugs)} drogas")
    print(f"  - {sum(len(d['lotes']) for d in output_drugs)} lotes (frascos)")
    print(f"  - {len(output['warehouses'])} depósitos")
    print(f"  - {len(output['users'])} usuarios")

    # Mostrar algunas estadísticas de clases
    from collections import Counter
    clase_counter = Counter()
    for d in output_drugs:
        for c in d["clases"]:
            clase_counter[c] += 1
    print("\nDistribución por clase:")
    for clase, count in clase_counter.most_common(10):
        print(f"  {clase}: {count}")


if __name__ == "__main__":
    main()
